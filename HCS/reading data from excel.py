import openpyxl
import statistics as st
deviations={} #space after or before = doesnot matter

wb = openpyxl.load_workbook('cricket_all.xlsx')
sheetnames= wb.sheetnames
# print sheetnames # gives the sheet names

for sheetname in sheetnames:
	sheet = wb[sheetname]
	strike_rates = []# create the list
	rows = sheet.max_row # Gives the max row count available
	
	for row in range(rows):
		runs = sheet.cell(row= row+1, column=1).value
		balls = sheet.cell(row= row+1, column=2).value
		strike_rate = runs/balls*100
		
		strike_rates.append(strike_rate)
	deviations[sheetname] = st.stdev(strike_rates)
	
for player in deviations:
	print(player, end='')
	print(deviations[player])
