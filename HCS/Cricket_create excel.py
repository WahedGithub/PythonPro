import openpyxl

wb = openpyxl.workbook.Workbook()# with this the Sheet will get created

sheet = wb["Sheet"] # this depends on the modules----and this we are defining to remove the sheet
#sheet = wb.get_sheet_by_name("Sheet") # this is used in older modules

wb.remove(sheet)
#wb.remove_sheet(sheet)

dravid = wb.create_sheet("dravid")
virat = wb.create_sheet("virat")

#sheet = wb.get_sheet_by_name("dravid")
sheet = wb["dravid"]
sheet.cell(row=1, column=1).value=100

#sheet = wb.get_sheet_by_name("Virat")
sheet = wb["virat"]
sheet.cell(row=1,column=1).value=200

wb.save("cricket.xlsx")
