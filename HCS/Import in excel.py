import openpyxl
filenames = ['dravid.txt', 'sachin.txt', 'virat.txt', 'yuvraj.txt']

wb = openpyxl.workbook.Workbook()
sheet = wb['Sheet']
wb.remove(sheet)

for filename in filenames:
	player= filename.replace(".txt", "")
	fobj = open(filename, 'r')
	sheet = wb.create_sheet(player)
	for line_number,line in enumerate(fobj):
		fields = line.split()
		runs = int(fields[2])
		balls = int(fields[3])
		
		sheet.cell(row=line_number+1, column=1).value=runs
		sheet.cell(row=line_number+1, column=2).value=balls


wb.save("cricket_all.xlsx")
