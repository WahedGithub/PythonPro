import openpyxl

wb = openpyxl.workbook.workbook()

#Sheet = wb["Sheet"]
sheet = wb.get_sheet_by_name("Sheet")

#wb.remove(sheet)
wb.remove_sheet(sheet)

dravid = wb.create_sheet("dravid")
Virat = wb.create_sheet("Virat")

sheet = wb.get_sheet_by_name("dravid")
sheet.cell(row=1, column=1).value=100

sheet = wb.get_sheet_by_name("Virat")
sheet.cell(row=1,column=1).value=200

wb.save("cricket.xlsx")
